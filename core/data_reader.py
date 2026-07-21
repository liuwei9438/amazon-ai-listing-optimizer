from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook

from .field_detector import detect_fields
from .models import WorkbookEnvelope
from .record_builder import build_product_records


def read_workbook(filename: str, raw_bytes: bytes) -> WorkbookEnvelope:
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("V2.2.2 当前只接受 .xlsx 文件。")

    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=False, data_only=False)
    visible_sheets = [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"]
    sheet_name = visible_sheets[0] if visible_sheets else workbook.sheetnames[0]

    dataframe = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=sheet_name, dtype=object).fillna("")
    fields, diagnostics = detect_fields(dataframe)
    records = build_product_records(dataframe, fields)

    ws = workbook[sheet_name]
    embedded_images = getattr(ws, "_images", [])
    embedded_details = []
    for image in embedded_images:
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is not None:
            embedded_details.append({"row": marker.row + 1, "column": marker.col + 1})

    diagnostics.update(
        {
            "sheet_name": sheet_name,
            "sheet_count": len(workbook.sheetnames),
            "embedded_image_count": len(embedded_images),
            "embedded_image_anchors": embedded_details,
            "merged_cell_ranges": len(ws.merged_cells.ranges),
            "has_formulas": any(
                isinstance(cell.value, str) and cell.value.startswith("=")
                for row in ws.iter_rows()
                for cell in row
            ),
            "record_count": len(records),
            "records_with_image_urls": sum(bool(record.image_urls) for record in records),
        }
    )
    workbook.close()

    return WorkbookEnvelope(
        filename=filename,
        raw_bytes=raw_bytes,
        dataframe=dataframe,
        sheet_name=sheet_name,
        fields=fields,
        records=records,
        diagnostics=diagnostics,
    )
