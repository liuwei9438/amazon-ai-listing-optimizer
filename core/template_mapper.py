
from __future__ import annotations

import io
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

TEMPLATE_HEADERS = [
    "父SKU(必填)", "SKU", "颜色", "分类", "中文简称", "英文简称",
    "库存", "币种", "成本价(必填)", "运费", "毛重(克)", "包装尺寸",
    "材料", "语言", "标题(必填)", "短标题", "关键字",
    "要点1", "要点2", "要点3", "要点4", "要点5",
    "简介", "产品图", "简介图", "参考网址",
]

ALIASES = {
    "颜色": ["颜色", "Color", "colour", "variant", "变体"],
    "分类": ["分类", "Category", "category_path", "类目"],
    "中文简称": ["中文简称", "中文名称", "中文名"],
    "英文简称": ["英文简称", "英文名称", "English Name"],
    "库存": ["库存", "Stock", "Quantity", "库存数量"],
    "币种": ["币种", "Currency"],
    "成本价(必填)": ["成本价(必填)", "成本价", "Cost", "Cost Price", "采购价"],
    "运费": ["运费", "Shipping", "Freight", "物流费"],
    "毛重(克)": ["毛重(克)", "毛重", "重量", "Weight", "Gross Weight"],
    "包装尺寸": ["包装尺寸", "Package Size", "Package Dimensions", "尺寸"],
    "材料": ["材料", "材质", "Material"],
    "简介图": ["简介图", "详情图", "Description Images", "A+ Images"],
    "参考网址": ["参考网址", "产品链接", "链接", "URL", "Reference URL"],
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _column_lookup(row: pd.Series) -> dict[str, str]:
    return {str(column).strip().casefold(): str(column) for column in row.index}


def _pick(row: pd.Series, aliases: list[str]) -> str:
    lookup = _column_lookup(row)
    for alias in aliases:
        column = lookup.get(alias.strip().casefold())
        if column is not None:
            value = _clean(row.get(column, ""))
            if value:
                return value
    return ""


def _has_column(row: pd.Series, name: str) -> bool:
    return name.strip().casefold() in _column_lookup(row)


def resolve_skus(source_row: pd.Series) -> tuple[str, str]:
    """
    Fixed business rule:
    - When source has `sku` and `child`, source sku is parent SKU and child is SKU.
    - When source has Parent SKU + SKU, Parent SKU is parent and SKU is child.
    - When only SKU exists, it is used as parent SKU and child SKU stays blank.
    """
    lookup = _column_lookup(source_row)

    child_column = next(
        (lookup[key] for key in ["child", "child sku", "child_sku", "子sku", "子 sku"]
         if key in lookup),
        None,
    )
    sku_column = lookup.get("sku")
    parent_column = next(
        (lookup[key] for key in ["父sku", "父 sku", "parent sku", "parent_sku", "parentsku"]
         if key in lookup),
        None,
    )

    if child_column is not None:
        return _clean(source_row.get(sku_column, "")) if sku_column else "", _clean(source_row.get(child_column, ""))

    if parent_column is not None:
        parent = _clean(source_row.get(parent_column, ""))
        child = _clean(source_row.get(sku_column, "")) if sku_column else ""
        return parent, child

    parent = _clean(source_row.get(sku_column, "")) if sku_column else ""
    return parent, ""


def map_results_to_template(
    result_df: pd.DataFrame,
    source_df: pd.DataFrame,
    *,
    title_col: str,
    short_title_col: str,
    bullet_cols: list[str],
    desc_col: str,
    image_col: str | None,
) -> pd.DataFrame:
    rows: list[dict[str, str]] = []

    for _, result_row in result_df.iterrows():
        if _clean(result_row.get("优化状态", "")) != "成功":
            continue

        source_index_raw = result_row.get("__源行索引", "")
        try:
            source_index = int(source_index_raw)
            source_row = source_df.loc[source_index]
        except Exception:
            source_row = result_row

        parent_sku, child_sku = resolve_skus(source_row)
        mapped = {header: "" for header in TEMPLATE_HEADERS}

        mapped["父SKU(必填)"] = parent_sku
        mapped["SKU"] = child_sku

        for target, aliases in ALIASES.items():
            mapped[target] = _pick(source_row, aliases)

        mapped["语言"] = _clean(result_row.get("__目标语言", result_row.get("语言", "")))
        mapped["标题(必填)"] = _clean(result_row.get(title_col, ""))
        mapped["短标题"] = _clean(result_row.get(short_title_col, ""))
        mapped["关键字"] = _clean(result_row.get("本地关键词", ""))

        for index, target in enumerate(["要点1", "要点2", "要点3", "要点4", "要点5"]):
            source_column = bullet_cols[index] if index < len(bullet_cols) else target
            mapped[target] = _clean(result_row.get(source_column, ""))

        mapped["简介"] = _clean(result_row.get(desc_col, ""))
        mapped["产品图"] = _clean(result_row.get(image_col, "")) if image_col else ""

        # If source has no explicit material but product analysis found a source-supported one,
        # use it. All other missing operational fields stay blank.
        if not mapped["材料"]:
            mapped["材料"] = _clean(result_row.get("识别材质", ""))

        rows.append(mapped)

    return pd.DataFrame(rows, columns=TEMPLATE_HEADERS)


def build_template_excel(
    mapped_df: pd.DataFrame,
    template_path: str | Path,
) -> bytes:
    template_path = Path(template_path)
    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    # Preserve the template header, widths, frozen panes and row-2 style.
    style_row = []
    if worksheet.max_row >= 2:
        for cell in worksheet[2]:
            style_row.append({
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            })

    if worksheet.max_row >= 2:
        worksheet.delete_rows(2, worksheet.max_row - 1)

    for row_index, values in enumerate(mapped_df.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if column_index - 1 < len(style_row):
                style = style_row[column_index - 1]
                cell.font = copy(style["font"])
                cell.fill = copy(style["fill"])
                cell.border = copy(style["border"])
                cell.alignment = copy(style["alignment"])
                cell.number_format = style["number_format"]
                cell.protection = copy(style["protection"])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
